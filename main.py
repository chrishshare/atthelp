# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import os
from shutil import copy
from pdUtil import *
from openpyxl import load_workbook
import time
import shutil


def replace_excel_title(excel, sheetname):
    workbook = load_workbook(excel)
    sheet = workbook[sheetname]
    sheet.cell(1, 1).value = 'name'
    sheet.cell(1, 2).value = 'company'
    sheet.cell(1, 3).value = 'service'
    sheet.cell(1, 4).value = 'atttime'
    sheet.cell(1, 5).value = 'terminal'
    workbook.save(excel)


def get_source_excel():
    files = os.listdir('.')
    for file in files:
        if file.endswith('.xlsx'):
            return file


def backup_source_excel(source):
    dst = source.split('.xlsx')[0] + '_备份.xlsx'
    copy(source, dst)


def backup_all_excel():
    dir = '结果文件_' + time.strftime("%Y%m%d%H%M%S", time.localtime())
    os.mkdir(dir)
    files = os.listdir('.')
    for file in files:
        if file.endswith('.xlsx'):
            shutil.move(file, dir)


if __name__ == '__main__':
    excel = get_source_excel()
    backup_source_excel(excel)
    src_sheetname = '原始打卡记录'
    replace_excel_title(excel=excel, sheetname=src_sheetname)
    res_excel = excel.split('.xlsx')[0] + '_结果.xlsx'

    read_excel_to_sqlite(excel=excel, sheetname='原始打卡记录')
    deal_to_sqlite()
    update_midnight()
    write_to_excel(excel=res_excel)

    backup_all_excel()
